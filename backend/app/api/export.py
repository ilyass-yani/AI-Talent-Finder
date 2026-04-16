"""Export endpoints for shortlist and candidate reports."""

from __future__ import annotations

import csv
import io
import json
from collections import defaultdict
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.dependencies import get_db
from app.models.models import Candidate, MatchResult, JobCriteria


router = APIRouter(prefix="/api/export", tags=["export"])


class ExportRequest(BaseModel):
    includeScores: bool = True
    includeSkills: bool = True
    includeExperience: bool = True
    includeEducation: bool = True
    sortBy: str = "score"
    criteria_id: Optional[int] = None


def _latest_criteria_id(db: Session) -> Optional[int]:
    criteria = db.query(JobCriteria).order_by(JobCriteria.created_at.desc()).first()
    return criteria.id if criteria else None


def _candidate_scores(db: Session, criteria_id: Optional[int]) -> dict[int, float]:
    if not criteria_id:
        criteria_id = _latest_criteria_id(db)
    if not criteria_id:
        return {}

    score_map = defaultdict(float)
    for match in db.query(MatchResult).filter(MatchResult.criteria_id == criteria_id).all():
        score_map[match.candidate_id] = match.score
    return score_map


def _sorted_candidates(candidates: List[Candidate], sort_by: str, score_map: dict[int, float]) -> List[Candidate]:
    if sort_by == "name":
        return sorted(candidates, key=lambda item: item.full_name.lower())
    if sort_by == "date":
        return sorted(candidates, key=lambda item: item.created_at, reverse=True)
    return sorted(candidates, key=lambda item: score_map.get(item.id, 0.0), reverse=True)


def _candidate_row(candidate: Candidate, score_map: dict[int, float], include_scores: bool, include_skills: bool, include_experience: bool, include_education: bool) -> List[str]:
    def _join_json_list(value: Optional[str]) -> str:
        if not value:
            return ""
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return "; ".join(str(item) for item in parsed)
        except Exception:
            pass
        return value

    row = [candidate.full_name, candidate.email, candidate.phone or ""]
    if include_scores:
        row.append(f"{score_map.get(candidate.id, 0.0):.2f}")
    if include_skills:
        skills = [skill.skill.name for skill in candidate.candidate_skills if skill.skill and skill.skill.name]
        row.append("; ".join(skills))
    if include_experience:
        row.append(_join_json_list(candidate.extracted_job_titles))
        row.append(_join_json_list(candidate.extracted_companies))
    if include_education:
        row.append(_join_json_list(candidate.extracted_education))
    return row


def _headers(settings: ExportRequest) -> List[str]:
    headers = ["Nom", "Email", "Téléphone"]
    if settings.includeScores:
        headers.append("Score")
    if settings.includeSkills:
        headers.append("Compétences")
    if settings.includeExperience:
        headers.extend(["Titres", "Compagnies"])
    if settings.includeEducation:
        headers.append("Éducation")
    return headers


def _rows(db: Session, settings: ExportRequest) -> List[List[str]]:
    score_map = _candidate_scores(db, settings.criteria_id)
    candidates = db.query(Candidate).all()
    ordered = _sorted_candidates(candidates, settings.sortBy, score_map)
    return [_headers(settings)] + [
        _candidate_row(candidate, score_map, settings.includeScores, settings.includeSkills, settings.includeExperience, settings.includeEducation)
        for candidate in ordered
    ]


def _csv_response(rows: List[List[str]], filename: str) -> StreamingResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerows(rows)
    payload = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))
    return StreamingResponse(payload, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _excel_response(rows: List[List[str]], filename: str) -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Candidates"

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)

    for row_index, row in enumerate(rows, start=1):
        sheet.append(row)
        if row_index == 1:
            for cell in sheet[row_index]:
                cell.fill = header_fill
                cell.font = header_font

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = str(cell.value or "")
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pdf_response(rows: List[List[str]], filename: str) -> StreamingResponse:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("1E3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("F8FAFC")]),
    ]))
    document.build([table])
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.post("/{export_format}")
def export_candidates(export_format: str, settings: ExportRequest, db: Session = Depends(get_db)):
    rows = _rows(db, settings)
    filename_root = f"candidates-export-{export_format}"

    if export_format == "csv":
        return _csv_response(rows, f"{filename_root}.csv")
    if export_format == "excel":
        return _excel_response(rows, f"{filename_root}.xlsx")
    if export_format == "pdf":
        return _pdf_response(rows, f"{filename_root}.pdf")

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported export format")
