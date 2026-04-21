"""
Export service — generate CSV / XLSX / PDF reports from candidate data.

Used by the `/api/export/{format}` route, but kept independent so other
contexts (CLI, scheduled job, email attachment) can reuse the same renderers
without dragging FastAPI in.
"""

from __future__ import annotations

import csv
import io
import json
from collections import defaultdict
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.models import Candidate, JobCriteria, MatchResult


# ---------------------------------------------------------------- Configuration

class ExportOptions:
    """Plain options bag — accepts both Pydantic models and raw kwargs."""

    def __init__(
        self,
        include_scores: bool = True,
        include_skills: bool = True,
        include_experience: bool = True,
        include_education: bool = True,
        sort_by: str = "score",
        criteria_id: Optional[int] = None,
    ) -> None:
        self.include_scores = include_scores
        self.include_skills = include_skills
        self.include_experience = include_experience
        self.include_education = include_education
        self.sort_by = sort_by
        self.criteria_id = criteria_id

    @classmethod
    def from_payload(cls, payload) -> "ExportOptions":
        """Accepts a Pydantic model with the legacy camelCase field names."""
        getter = (lambda key, default: getattr(payload, key, default)) if hasattr(payload, "__dict__") else (lambda key, default: payload.get(key, default))
        return cls(
            include_scores=getter("includeScores", True),
            include_skills=getter("includeSkills", True),
            include_experience=getter("includeExperience", True),
            include_education=getter("includeEducation", True),
            sort_by=getter("sortBy", "score"),
            criteria_id=getter("criteria_id", None),
        )


# ---------------------------------------------------------------- Data assembly

def _latest_criteria_id(db: Session) -> Optional[int]:
    criteria = db.query(JobCriteria).order_by(JobCriteria.created_at.desc()).first()
    return criteria.id if criteria else None


def _candidate_scores(db: Session, criteria_id: Optional[int]) -> dict[int, float]:
    if not criteria_id:
        criteria_id = _latest_criteria_id(db)
    if not criteria_id:
        return {}

    score_map: dict[int, float] = defaultdict(float)
    for match in db.query(MatchResult).filter(MatchResult.criteria_id == criteria_id).all():
        score_map[match.candidate_id] = match.score
    return score_map


def _sorted_candidates(candidates: Iterable[Candidate], sort_by: str, score_map: dict[int, float]) -> List[Candidate]:
    candidates = list(candidates)
    if sort_by == "name":
        return sorted(candidates, key=lambda item: (item.full_name or "").lower())
    if sort_by == "date":
        return sorted(candidates, key=lambda item: item.created_at, reverse=True)
    return sorted(candidates, key=lambda item: score_map.get(item.id, 0.0), reverse=True)


def _join_json_list(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            return "; ".join(str(item) for item in parsed)
    except Exception:
        pass
    return value


def _candidate_row(candidate: Candidate, score_map: dict[int, float], options: ExportOptions) -> List[str]:
    row: List[str] = [candidate.full_name or "", candidate.email or "", candidate.phone or ""]
    if options.include_scores:
        row.append(f"{score_map.get(candidate.id, 0.0):.2f}")
    if options.include_skills:
        skills = [cs.skill.name for cs in candidate.candidate_skills if cs.skill and cs.skill.name]
        row.append("; ".join(skills))
    if options.include_experience:
        row.append(_join_json_list(candidate.extracted_job_titles))
        row.append(_join_json_list(candidate.extracted_companies))
    if options.include_education:
        row.append(_join_json_list(candidate.extracted_education))
    return row


def _headers(options: ExportOptions) -> List[str]:
    headers = ["Nom", "Email", "Téléphone"]
    if options.include_scores:
        headers.append("Score")
    if options.include_skills:
        headers.append("Compétences")
    if options.include_experience:
        headers.extend(["Titres", "Compagnies"])
    if options.include_education:
        headers.append("Éducation")
    return headers


def build_rows(db: Session, options: ExportOptions) -> List[List[str]]:
    """Return the full table (header row + candidate rows) for an export."""
    score_map = _candidate_scores(db, options.criteria_id)
    candidates = db.query(Candidate).all()
    ordered = _sorted_candidates(candidates, options.sort_by, score_map)
    return [_headers(options)] + [_candidate_row(candidate, score_map, options) for candidate in ordered]


# ---------------------------------------------------------------- Renderers

def render_csv(rows: List[List[str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerows(rows)
    # BOM helps Excel open UTF-8 CSV with proper accents.
    return buffer.getvalue().encode("utf-8-sig")


def render_xlsx(rows: List[List[str]], sheet_title: str = "Candidates") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)

    for row_index, row in enumerate(rows, start=1):
        sheet.append(row)
        if row_index == 1:
            for cell in sheet[row_index]:
                cell.fill = header_fill
                cell.font = header_font

    for column_cells in sheet.columns:
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_pdf(rows: List[List[str]]) -> bytes:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    document.build([table])
    return buffer.getvalue()


SUPPORTED_FORMATS = {"csv", "excel", "pdf"}


def export_candidates(db: Session, fmt: str, options: ExportOptions) -> tuple[bytes, str, str]:
    """High-level entry point: returns (payload_bytes, media_type, filename)."""
    fmt = fmt.lower()
    if fmt not in SUPPORTED_FORMATS:
        raise ValueError(f"Unsupported export format: {fmt}")

    rows = build_rows(db, options)
    base = f"candidates-export-{fmt}"

    if fmt == "csv":
        return render_csv(rows), "text/csv", f"{base}.csv"
    if fmt == "excel":
        return (
            render_xlsx(rows),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{base}.xlsx",
        )
    return render_pdf(rows), "application/pdf", f"{base}.pdf"
